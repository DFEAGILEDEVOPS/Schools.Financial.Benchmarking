import pandas as pd
import numpy as np
import openpyxl
import cfr_data_mapping
import federations_mapping

# Expected header row in `input.xlsx``:
# LAEstab	LAEstabOld	LA	LAOld	LA Name	Estab	URN	SchoolName	PeriodCoveredByReturn	DidNotSupplyFlag	FederationYN	LeadSchool	General Hospital School Indicator	IndividualPupilsFTE	NoPupils	IndTeachers_FTE	NoTeachers	Gender	OverallPhase	Phase	Type	UrbanRural	Region	LondonBorough	LondonWeighting	Ind_PC_FSM	PercentageOfPupilsEligibleForFSM	Ind_PC_EHCP	PercentageOfSENPupilsWithoutStatement	Ind_PC_SEN_Support	PercentageOfPupilsWithSENStatement	Ind_PC_EAL	PercentageOfPupilsWithEAL	Ind_PC_Boarders	PercentageOfPupilsWhoAreBoarders	AdmissionsPolicy	PFI	HasA6thForm	NoOfPupilsIn6thForm	Aggregated_VIthForm	LowestAgeOfPupils	HighestAgeOfPupils	Ind_TA_FTE	FTE of Teaching Assistants	% of teachers with QTS	Pre16Funding	Post16Funding	I01 2 Pre and Post-16 Funding	SEN	FundingForMinorityEthnicPupils	PupilPremium	OtherDfEEFARevenueGrants	OtherIncomeLAandOtherGovernmentGrants	I08a Income from lettings	I08b Income from facilities and services	IncomeFromFacilitiesAndServices	IncomeFromCatering	ReceiptsFromSupplyTeacherInsuranceClaims	ReceiptsFromOtherInsuranceClaims	IncomeFromContributionsToVisitsEtc	DonationsAndOrVoluntaryFunds	PupilFocussedExtendedSchoolFundingAndOrGrants	CommunityFocussedSchoolFundingAndOrGrants	CommunityFocusedSchoolFacilitiesIncome	I18a Income from the Coronavirus Job Retention Scheme	I18b Income from the DfE grant scheme for reimbursing exceptional costs associated with COVID-19	I18c Income from the Â£1bn COVID-19 catch-up package announced on 20 July 2020	I18d Income from other additional grants	AdditionalGrantForSchools	Total Income  I01 to I08, I11 to I15, I18 minus E30	Total Income  I01 to I18 minus E30	TeachingStaff	SupplyTeachingStaff	EducationSupportStaff	PremisesStaff	AdministrativeAndClericalStaff	CateringStaff	OtherStaff	IndirectEmployeeExpenses	StaffDevelopmentAndTraining	SupplyTeacherInsurance	StaffRelatedInsurance	BuildingMaintenanceAndImprovement	GroundsMaintenanceAndImprovement	CleaningAndCaretaking	WaterAndSewerage	Energy	Rates	OtherOccupationCosts	LearningResourcesNotICTEquipment	ICTLearningResources	ExaminationFees	AdministrativeSuppliesNonEducational	OtherInsurancePremiums	SpecialFacilities	CateringSupplies	AgencySupplyTeachingStaff	EducationalConsultancy	E28a Bought in professional services, other, not PFI	PFICharges	BoughtInProfessionalServicesOther	InterestChargesForLoansAndBank	DirectRevenueFinancingRevenueContributionsToCapital	CommunityFocusedSchoolStaff	CommunityFocusedSchoolCosts	Total Expenditure E01 to E29 and E31 to E32 minus I9, I10, I16 and I17	Total Expenditure E01 to E29 and E31 to E32	OB01 Opening pupil-focused revenue balance	OB02 Opening community-focused revenue balance	OB03 Opening capital balance	CI01 Capital income	CI03 Voluntary or private income	DirectRevenueFinancingCapitalReserveTransfers	CE01 Acquisition of land and existing buildings	CE02 New construction, conversion and renovation	CE03 Vehicles, plant, equipment and machinery	CE04 Information and communication technology	B01 Committed revenue balances	B02 Uncommitted revenue balances	B03 Devolved formula capital balance	B05 Other capital balances	B06 Community-focused school revenue balances	B07 Outstanding balance on capital loans to school	RevenueReserve	InYearBalance	GrantFunding	DirectGrant	CommunityGrants	TargetedGrants	SelfGeneratedFunding	TotalIncome	Teaching Staff  E01	SupplyStaff	Education support staff  E03	OtherStaffCosts	StaffTotal	MaintenanceAndImprovement	Premises	CateringExp	Occupation	SuppliesAndServices	EducationalSupplies	BroughtInProfessionalSevices	CommunityExp	TotalExpenditure	FTE of Support Staff	FTE of Admin Staff
# Also ensure that the column order has not changed when C&P-ing the above header row...

def create_output_file(input_file, output_file):
    
    df = pd.read_excel(input_file)

    empty_df = pd.DataFrame()

    # df for each sheet
    federations_df = create_federations_df(df)
    cfr_data_df = create_cfr_data_df(df)

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        empty_df.to_excel(writer, sheet_name='Guidance', index=False)
        empty_df.to_excel(writer, sheet_name='CFR Groupings', index=False)
        federations_df.to_excel(writer, sheet_name='Federations', index=False)
        cfr_data_df.to_excel(writer, sheet_name='CFR_Data', index=False)


def create_federations_df(input_df):   
    # Filter df to include only federations and sort into groups
    is_federation = ~(input_df["FederationYN"] == "No")
    filtered_df = input_df[is_federation]   
    sorted_df = filtered_df.sort_values(by=["LeadSchool", "FederationYN"])
    
    # new df with only required columns
    new_df = sorted_df[federations_mapping.columns_to_include].copy()

    # insert new columns for calcs
    new_df.insert(3, "Total expenditure per pupil", np.nan)
    new_df.insert(7, "FTE pupils of federation", np.nan)

    # rename columns based on mapping
    new_df = new_df.rename(columns=federations_mapping.rename_col_dict)

    # Populate the FTE pupils of federation column only where lead school
    new_df['FTE pupils of federation'] = new_df.apply(
        lambda row: new_df[new_df['Lead school'] == row['Lead school']]['FTE pupils of individual schools'].sum() if row['LAESTAB'] == row['Lead school'] else None,
        axis=1
    )

    # Populate the Total expenditure per pupil column
    new_df['Total expenditure per pupil'] = new_df['Total Expenditure calc only DELETE'] / new_df['FTE pupils of federation']
    # drop column no longer required
    new_df.drop(['Total Expenditure calc only DELETE'], axis=1, inplace=True)

    # insert empty row to seperate each federated group
    result_df = pd.DataFrame(columns=new_df.columns)
    empty_row = pd.DataFrame([[np.nan] * len(new_df.columns)], columns=new_df.columns)
    last_lead_school = None
    for _, row in new_df.iterrows():
        if row['Lead school'] != last_lead_school:
            result_df = pd.concat([result_df, empty_row], ignore_index=True)

        result_df = pd.concat([result_df, row.to_frame().T], ignore_index=True)

        

        last_lead_school = row['Lead school']

    result_df = pd.concat([result_df, row.to_frame().T], ignore_index=True)

    return result_df


def create_cfr_data_df(input_df):   
    # new df with only required columns
    new_df = input_df[cfr_data_mapping.columns_to_include].copy()
    
    # insert col for calc
    new_df.insert(14, "Full time equivalent number of pupils in federation", np.nan)

    # rename cols based on mapping
    result_df = new_df.rename(columns=cfr_data_mapping.rename_col_dict)

    # sum of total pupil in federation for lead school - for others == total pupil in school 
    result_df['Full time equivalent number of pupils in federation'] = result_df.apply(
        lambda row: result_df[result_df['Lead school in federation'] == row['Lead school in federation']]['Full time equivalent number of pupils in school']
        .sum() if row['LAEstab'] == row['Lead school in federation'] else row['Full time equivalent number of pupils in school'],
        axis=1
    )

    return result_df


def autosize_columns(sheet):
    for col in sheet.columns:
        max_length = 0
        column = [cell for cell in col]
        for cell in column:
            try:  
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width


def insert_title_cells_and_format(file):
    wb = openpyxl.load_workbook(file)

    # sheets in workbook to be edited
    federations_sheet = wb["Federations"]
    cfr_data_sheet = wb["CFR_Data"]

    autosize_columns(federations_sheet)
    autosize_columns(cfr_data_sheet)

    federations_sheet.insert_rows(0, amount=3)
    
    # set values and format for title
    federations_sheet_title_value = "Federations in Consistent Financial Reporting 2022-23"
    federations_sheet_title_cell = federations_sheet.cell(row=1, column=1)
    federations_sheet_title_cell.value=federations_sheet_title_value
    federations_sheet_title_cell.font = openpyxl.styles.Font(bold=True) 

    # set values and format for description
    federations_sheet_description_value = "A number of schools are federated together, providing combined income and expenditure data. Where this is the case, the figures given in the other sheets in this workbook represent the whole federation. This sheet details which schools are grouped in federations, and how many pupils they contribute to the total for the federation."
    federations_sheet_description_cell = federations_sheet.cell(row=2, column=1)
    federations_sheet_description_cell.value=federations_sheet_description_value
    federations_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    federations_sheet_description_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    federations_sheet.row_dimensions[2].height = 50
    

    # format Total expenditure per pupil col as currency
    column_letter = openpyxl.utils.get_column_letter(4)
    for cell in federations_sheet[column_letter]:
        cell.number_format= '"£"#,##0'


    cfr_data_sheet.insert_rows(0, amount=2)

    # set values and format for title
    cfr_data_sheet_title_value = "School level total income and expenditure data 2022-23 (£)"
    cfr_data_sheet_title_value_cell = cfr_data_sheet.cell(row=1, column=1)
    cfr_data_sheet_title_value_cell.value=cfr_data_sheet_title_value
    cfr_data_sheet_title_value_cell.font = openpyxl.styles.Font(bold=True) 

    wb.save('output.xlsx')


if __name__ == "__main__":
    create_output_file("input.xlsx", "output.xlsx")
    insert_title_cells_and_format("output.xlsx")
